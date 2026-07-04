import csv
import json
from pathlib import Path


class ExtractionError(RuntimeError):
    pass


def _truncate(text: str, limit: int) -> str:
    if len(text) <= limit:
        return text
    head = int(limit * 0.7)
    tail = limit - head
    return f"{text[:head]}\n\n...（中略）...\n\n{text[-tail:]}"


def extract_text(path: Path, max_length: int = 30_000) -> str:
    suffix = path.suffix.lower()
    try:
        if suffix in {".txt", ".md"}:
            text = path.read_text(encoding="utf-8", errors="replace")
        elif suffix == ".json":
            data = json.loads(path.read_text(encoding="utf-8", errors="replace"))
            text = json.dumps(data, ensure_ascii=False, indent=2)
        elif suffix == ".csv":
            rows = []
            with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
                for idx, row in enumerate(csv.reader(handle)):
                    rows.append(" | ".join(row))
                    if idx >= 500:
                        break
            text = "\n".join(rows)
        elif suffix == ".pdf":
            import fitz
            doc = fitz.open(path)
            text = "\n".join(page.get_text("text") for page in doc)
        elif suffix == ".docx":
            from docx import Document
            doc = Document(path)
            text = "\n".join(paragraph.text for paragraph in doc.paragraphs)
        elif suffix == ".xlsx":
            from openpyxl import load_workbook
            workbook = load_workbook(path, read_only=True, data_only=True)
            rows = []
            for sheet in workbook.worksheets[:10]:
                rows.append(f"# Sheet: {sheet.title}")
                for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    rows.append(" | ".join("" if value is None else str(value) for value in row))
                    if idx >= 300:
                        break
            text = "\n".join(rows)
        else:
            return ""
    except Exception as exc:
        raise ExtractionError(f"本文抽出に失敗しました: {path.name}") from exc
    return _truncate(text.strip(), max_length)
